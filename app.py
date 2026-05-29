import re
import zipfile
from io import BytesIO
from typing import Dict, Tuple

import streamlit as st
from openpyxl import load_workbook, Workbook

st.set_page_config(
    page_title="Excel 關鍵字替換工具",
    page_icon="🧰",
    layout="centered",
)

DEFAULT_RULES = """蝦皮=>
聊聊=>
shopee=>
Shopee=>
SHOPEE=>"""

TRACKING_PARAMS = [
    "utm_source", "utm_medium", "utm_campaign", "utm_term", "utm_content",
    "fbclid", "gclid", "yclid", "mc_cid", "mc_eid"
]

# XML 1.0 不允許的控制字元，很多平台匯出的 xlsx 會混入這些字元。
INVALID_XML_RE = re.compile(
    r"[\x00-\x08\x0B\x0C\x0E-\x1F\x7F]"
)


def parse_rules(text: str, ignore_case: bool) -> Dict[str, str]:
    rules: Dict[str, str] = {}
    for line in text.splitlines():
        line = line.strip()
        if not line or line.startswith("#"):
            continue
        if "=>" not in line:
            continue
        old, new = line.split("=>", 1)
        old = old.strip()
        new = new.strip()
        if old:
            rules[old] = new
    return rules


def replace_text(value: str, rules: Dict[str, str], ignore_case: bool) -> str:
    result = value
    for old, new in rules.items():
        if ignore_case:
            result = re.sub(re.escape(old), new, result, flags=re.IGNORECASE)
        else:
            result = result.replace(old, new)
    return result


def remove_tracking_params(url: str) -> str:
    if "?" not in url:
        return url

    base, query = url.split("?", 1)
    fragment = ""
    if "#" in query:
        query, fragment = query.split("#", 1)
        fragment = "#" + fragment

    kept = []
    for part in query.split("&"):
        if not part:
            continue
        key = part.split("=", 1)[0]
        if key not in TRACKING_PARAMS:
            kept.append(part)

    if kept:
        return base + "?" + "&".join(kept) + fragment
    return base + fragment


def clean_xml_bytes(data: bytes) -> bytes:
    """移除常見非法 XML 字元，讓 openpyxl 比較容易讀取髒 xlsx。"""
    try:
        text = data.decode("utf-8")
    except UnicodeDecodeError:
        try:
            text = data.decode("utf-8", errors="ignore")
        except Exception:
            return data
    text = INVALID_XML_RE.sub("", text)
    return text.encode("utf-8")


def repair_xlsx(uploaded_file) -> BytesIO:
    """把 xlsx 視為 zip，清洗內部 XML 後重新打包。"""
    uploaded_file.seek(0)
    original = BytesIO(uploaded_file.read())
    repaired = BytesIO()

    with zipfile.ZipFile(original, "r") as zin, zipfile.ZipFile(repaired, "w", zipfile.ZIP_DEFLATED) as zout:
        for item in zin.infolist():
            data = zin.read(item.filename)
            if item.filename.lower().endswith((".xml", ".rels")):
                data = clean_xml_bytes(data)
            zout.writestr(item, data)

    repaired.seek(0)
    return repaired


def load_workbook_with_auto_repair(uploaded_file) -> Tuple[Workbook, str]:
    """先正常讀取；失敗就自動修復 XML 後再讀。"""
    uploaded_file.seek(0)
    try:
        wb = load_workbook(uploaded_file)
        return wb, "正常讀取"
    except Exception as first_error:
        try:
            repaired_file = repair_xlsx(uploaded_file)
            wb = load_workbook(repaired_file)
            return wb, f"已啟用自動修復模式：{first_error}"
        except Exception as second_error:
            raise RuntimeError(
                "檔案內部結構損壞太嚴重，系統無法自動修復。"
                "請確認檔案是真正的 .xlsx，或先用 Excel 開啟後另存一次。"
                f" 原始錯誤：{first_error}；修復後錯誤：{second_error}"
            )


def process_excel(uploaded_file, rules: Dict[str, str], ignore_case: bool, clean_tracking: bool) -> Tuple[BytesIO, int, int, str]:
    wb, load_mode = load_workbook_with_auto_repair(uploaded_file)
    changed_cells = 0
    changed_links = 0

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    original = cell.value
                    updated = replace_text(original, rules, ignore_case)
                    if clean_tracking and updated.startswith(("http://", "https://")):
                        updated = remove_tracking_params(updated)
                    if updated != original:
                        cell.value = updated
                        changed_cells += 1

                if cell.hyperlink and cell.hyperlink.target:
                    original_target = cell.hyperlink.target
                    updated_target = replace_text(original_target, rules, ignore_case)
                    if clean_tracking:
                        updated_target = remove_tracking_params(updated_target)
                    if updated_target != original_target:
                        cell.hyperlink = updated_target
                        changed_links += 1

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output, changed_cells, changed_links, load_mode


def safe_download_name(filename: str) -> str:
    """下載檔名維持上傳檔名，不另外加 processed 或日期。"""
    if not filename:
        return "download.xlsx"
    return filename


st.title("Excel 關鍵字替換工具")
st.caption("上傳 Excel，批次替換儲存格文字與超連結網址。已加入髒檔自動修復與原檔名下載。")

with st.expander("使用方式", expanded=True):
    st.write("每行一組替換規則，格式為 `原文字=>新文字`。如果右邊留空，就代表刪除該關鍵字。")
    st.code("蝦皮=>\n聊聊=>\nshopee=>平台", language="text")

uploaded_file = st.file_uploader("上傳 Excel 檔案", type=["xlsx"])

rules_text = st.text_area(
    "替換規則",
    value=DEFAULT_RULES,
    height=180,
    help="範例：蝦皮=>平台。右側留空代表刪除。",
)

col1, col2 = st.columns(2)
with col1:
    ignore_case = st.checkbox("英文大小寫不區分", value=True)
with col2:
    clean_tracking = st.checkbox("順便移除常見追蹤參數", value=True)

if st.button("開始處理", type="primary", use_container_width=True):
    if uploaded_file is None:
        st.error("請先上傳 Excel 檔案。")
    else:
        rules = parse_rules(rules_text, ignore_case)
        if not rules and not clean_tracking:
            st.error("請至少輸入一組替換規則，或勾選移除追蹤參數。")
        else:
            try:
                output, changed_cells, changed_links, load_mode = process_excel(
                    uploaded_file, rules, ignore_case, clean_tracking
                )
                if load_mode == "正常讀取":
                    st.info("檔案讀取狀態：正常讀取。")
                else:
                    st.warning("檔案讀取狀態：已自動修復髒檔後處理。部分損壞樣式可能會被 Excel 重新整理。")
                st.success(f"處理完成：已修改 {changed_cells} 個儲存格、{changed_links} 個超連結。")
                st.download_button(
                    label="下載處理後 Excel",
                    data=output,
                    file_name=safe_download_name(uploaded_file.name),
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            except Exception as exc:
                st.error(f"處理失敗：{exc}")

st.divider()
st.caption("提醒：此工具會在記憶體中處理檔案，不會主動保存上傳內容。下載檔名會維持與上傳檔案相同；若本機已有同名檔，瀏覽器可能自動加上 (1)。")
